"""
PDF Financial Table Extractor — Final Version
Uses header-anchored column detection for accurate parsing across different layouts.
"""
import re, os
from pathlib import Path
from collections import defaultdict
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# Filenames the pipeline must NEVER overwrite. Adding more here is the
# project-wide allow-listing mechanism. Match by basename, case-insensitive,
# wildcard-extension.
_PROTECTED_BASENAMES = {"input templates"}


def _assert_not_protected(path) -> None:
    """Raise if writing to `path` would clobber a manually-curated source
    template (e.g. the root `Input Templates.xlsx`). The pipeline never
    needs to write these — any code path that does is a bug."""
    p = Path(path)
    stem = p.stem.lower().strip()
    if stem in _PROTECTED_BASENAMES:
        raise PermissionError(
            f"Refusing to write to protected file: {p}. "
            f"This file is a hand-curated source template and must not be "
            f"modified by the pipeline.")

STOP_PHRASES = [
    'summary of', 'accounting polic', 'accompanying notes', 'as per our report',
    'chartered accountant', 'firm registration', 'icai firm', 'membership no',
    'place:', 'date:', 'sd/-', 'din:', 'pan:', 'partner', ' director',
    'company secretary', 'chief financial', 'whole time',
    'on behalf of', 'for and on behalf', 'cin:', 'material accounting'
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Unit detection — normalize all extracted values to LAKHS (canonical)
# so the mapper, per-company workbook, and Airtable upload speak the same
# scale across reports. Without this, a Crore-denominated FY24-25 report
# would compare 100× wrong against a Lakh-denominated FY23-24 in the
# per-company workbook.
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Multipliers to convert a value to LAKHS.
# Crore = 100 lakhs; Million = 10 lakhs; Thousand = 0.01 lakhs; Rupee = 1e-5
_UNIT_MULT_TO_LAKHS = {
    "lakh":     1.0,
    "crore":    100.0,
    "million":  10.0,
    "billion":  10000.0,
    "thousand": 0.01,
    "rupee":    1e-5,
}

# Patterns to detect the declared unit. Order matters — more specific first.
# All matched against the page text (lowercased, whitespace-collapsed).
_UNIT_PATTERNS = [
    (r"\b(?:in\s+)?(?:rs\.?|inr|₹)?\s*(?:in\s+)?crore[s]?\b", "crore"),
    (r"\b(?:in\s+)?(?:rs\.?|inr|₹)?\s*(?:in\s+)?lakh[s]?\b",  "lakh"),
    (r"\b(?:in\s+)?(?:usd|inr|₹)?\s*(?:in\s+)?million[s]?\b", "million"),
    (r"\b(?:in\s+)?(?:usd|inr|₹)?\s*(?:in\s+)?billion[s]?\b", "billion"),
    (r"\b(?:in\s+)?(?:rs\.?|inr|₹)?\s*(?:in\s+)?thousand[s]?\b", "thousand"),
]


def detect_unit_from_text(text: str) -> str:
    """Return one of: 'crore', 'lakh', 'million', 'billion', 'thousand',
    or 'lakh' as the safe default for Indian Schedule-III statements when
    the unit is genuinely absent from the header."""
    if not text:
        return "lakh"
    flat = re.sub(r"\s+", " ", text.lower())[:1500]
    for pat, unit in _UNIT_PATTERNS:
        if re.search(pat, flat):
            return unit
    return "lakh"


def normalize_rows_to_lakhs(rows, source_unit: str):
    """Multiply every numeric cy/py in `rows` by the conversion factor that
    maps `source_unit` → LAKHS. Returns the (mutated) rows list and the
    factor that was applied."""
    factor = _UNIT_MULT_TO_LAKHS.get(source_unit, 1.0)
    if factor == 1.0:
        return rows, factor
    for r in rows:
        for k in ("cy", "py"):
            v = r.get(k)
            if isinstance(v, (int, float)):
                r[k] = round(float(v) * factor, 2)
    return rows, factor


# Single-word page-design / cover-strip text that bleeds into the rightmost
# data column on annual reports. These are layout / watermark tokens, not
# financial data — drop the row if a "value" column contains only one of
# these (the rest of the row is empty, see Ambuja FY22-23).
WATERMARK_TOKENS = {
    'creating', 'value', 'creation', 'accelerating', 'our', 'growth',
    'sustainability', 'statutory', 'reports', 'report', 'financial',
    'statements', 'annual', 'corporate', 'overview', 'governance',
    'integrated', 'highlights', 'standalone', 'consolidated',
    'building', 'beyond', 'strength', 'progress', 'continued',
}

STMT_KEYWORDS = {
    'consolidated_pl': ['CONSOLIDATED', 'PROFIT'],
    'standalone_pl': ['STANDALONE', 'PROFIT'],
    'consolidated_bs': ['CONSOLIDATED', 'BALANCE SHEET'],
    'standalone_bs': ['STANDALONE', 'BALANCE SHEET'],
    'generic_bs': ['BALANCE SHEET'],
    'generic_pl': ['PROFIT'],
}

STMT_NAMES = {
    'consolidated_pl': 'Consolidated P&L',
    'standalone_pl': 'Standalone P&L',
    'consolidated_bs': 'Consolidated Balance Sheet',
    'standalone_bs': 'Standalone Balance Sheet',
    'generic_bs': 'Balance Sheet',
    'generic_pl': 'P&L',
}


def classify_page(text):
    upper = text[:500].upper()
    for key, keywords in STMT_KEYWORDS.items():
        if all(k in upper for k in keywords):
            return key
    return None


def _strip_sidebar_watermarks(words, page_width):
    """Remove design-strip text that lives in the extreme right margin of
    annual reports (e.g. Ambuja's "Creating / Value / Statutory / Reports"
    sidebar at x≈568 on a 595pt-wide page).

    These words pollute column detection because they appear EARLIER than
    the body text in pdfplumber's iteration order — any loop that bounds
    by `top` will hit them first and miss the actual table headers.

    Heuristic: a word is sidebar text if it's a single design token
    (from WATERMARK_TOKENS) AND sits in the rightmost 5% of the page.
    """
    right_strip = page_width * 0.94
    return [w for w in words
            if not (w['x0'] >= right_strip
                    and w['text'].strip().lower() in WATERMARK_TOKENS)]


def find_column_boundaries(words, page_width):
    """
    Anchor columns from header words: 'Note'/'Notes' and year date columns.
    """
    note_word = None
    year_words = []

    # Scan ALL header-zone words (not `break` on first top>limit — some PDFs
    # emit sidebar/watermark words at top>200 BEFORE the actual table header
    # words in iteration order, which would skip the real headers entirely).
    note_candidates = []
    for w in words:
        if w['top'] > 250:
            continue
        txt = w['text'].strip()
        low = txt.lower()
        # Collect ALL "Note"/"Notes" words then pick the leftmost. Indian
        # annual reports often have a parenthetical "(Refer Note - 69)" in
        # the header row sitting RIGHT of the actual "Notes" column header;
        # the real one is always the leftmost.
        if low in ('note', 'notes'):
            note_candidates.append(w)
        # Year header words. Lowered the x0 threshold from 0.55 to 0.40 —
        # narrower-margin layouts (Ambuja consolidated BS) put the CY column
        # at x≈438 on a 595pt page, just under the old 0.55 cutoff.
        if re.match(r'^20\d{2}$', txt) and w['x0'] > page_width * 0.40:
            year_words.append(w)

    if note_candidates:
        # Prefer "Notes" (plural — column header is usually plural) over
        # "Note" (often part of a "Refer Note - X" caption). Tie-break by x0.
        note_candidates.sort(key=lambda w: (0 if w['text'].strip().lower() == 'notes' else 1,
                                             w['x0']))
        note_word = note_candidates[0]
    
    cy_center = None
    py_center = None
    
    if len(year_words) >= 2:
        year_words.sort(key=lambda w: w['x0'])
        cy_center = (year_words[0]['x0'] + year_words[0]['x1']) / 2
        py_center = (year_words[-1]['x0'] + year_words[-1]['x1']) / 2
    
    if cy_center is None or py_center is None:
        return _fallback_column_detection(words, page_width, note_word)
    
    mid_cy_py = (cy_center + py_center) / 2
    
    # Find actual note number positions and first CY value positions from data
    note_data_rights = []
    cy_data_lefts = []
    for w in words:
        if w['top'] < 100:
            continue
        txt = w['text'].strip()
        x0 = w['x0']
        # Note numbers: 1-2 digit integers in the note column area
        if (re.match(r'^\d{1,2}$', txt) and int(txt) <= 40
                and note_word and abs(x0 - note_word['x0']) < 30):
            note_data_rights.append(w['x1'])
        # CY values: numeric values left of midpoint
        elif (re.match(r'^[\d,(.)]+$', txt.replace(' ', '')) 
              and x0 > (note_word['x1'] + 10 if note_word else page_width * 0.5)
              and (x0 + w['x1'])/2 < mid_cy_py):
            cy_data_lefts.append(x0)
    
    if note_word:
        note_x0 = note_word['x0']
        # Use actual note data extent if available, otherwise header + small margin
        if note_data_rights:
            note_right = max(note_data_rights) + 8
        else:
            note_right = note_word['x1'] + 15
    else:
        note_x0 = cy_center - 60
        note_right = cy_center - 40
    
    # CY left: use actual leftmost CY value, or note_right + gap
    if cy_data_lefts:
        cy_left = min(cy_data_lefts) - 5
    else:
        cy_left = note_right + 5
    
    # Find actual value boundaries from data
    # CY values should end near cy_center + some offset
    cy_val_rights = []
    py_val_rights = []
    for w in words:
        if w['top'] < 100:
            continue
        txt = w['text'].strip().replace(',', '').replace('(', '').replace(')', '').replace(' ', '')
        if not re.match(r'^[\d.]+$', txt) or txt.count('.') > 1:
            continue
        x_mid = (w['x0'] + w['x1']) / 2
        if x_mid < mid_cy_py and w['x0'] > cy_left - 10:
            cy_val_rights.append(w['x1'])
        elif x_mid > mid_cy_py:
            py_val_rights.append(w['x1'])
    
    return {
        'particulars_right': note_x0 - 5 if note_word else cy_left - 30,
        'note_left': note_x0 - 5 if note_word else cy_left - 30,
        'note_right': note_right,
        'cy_left': cy_left,
        'cy_right': mid_cy_py,
        'py_left': mid_cy_py,
        'py_right': max(py_val_rights) + 20 if py_val_rights else page_width,
    }


def _fallback_column_detection(words, page_width, note_word):
    """Fallback when year headers aren't found."""
    num_positions = []
    for w in words:
        if w['top'] < 100:
            continue
        txt = w['text'].strip().replace(',', '').replace('(', '').replace(')', '').replace(' ', '')
        if re.match(r'^[\d,]+\.?\d{0,2}$', txt) and txt.count('.') <= 1:
            try:
                fval = float(txt.replace(',', ''))
            except ValueError:
                continue
            if fval > 0.001:
                num_positions.append((w['x0'], w['x1']))
    
    if len(num_positions) < 4:
        return None
    
    # Cluster right edges
    rights = sorted(set(round(p[1], 0) for p in num_positions))
    clusters = []
    cur = [rights[-1]]
    for x in reversed(rights[:-1]):
        if cur[-1] - x < 20:
            cur.append(x)
        else:
            clusters.append(cur)
            cur = [x]
            if len(clusters) >= 3:
                break
    clusters.append(cur)
    
    if len(clusters) < 2:
        return None
    
    py_right = max(clusters[0])
    cy_right = max(clusters[1])
    
    cy_lefts = [p[0] for p in num_positions if abs(p[1] - cy_right) < 20]
    py_lefts = [p[0] for p in num_positions if abs(p[1] - py_right) < 20]
    cy_left = min(cy_lefts) if cy_lefts else cy_right - 80
    py_left = min(py_lefts) if py_lefts else py_right - 80
    mid_cy_py = (cy_right + py_left) / 2
    
    note_x0 = note_word['x0'] if note_word else cy_left - 30
    note_right = (note_word['x1'] + 20) if note_word else cy_left - 5
    
    return {
        'particulars_right': note_x0 - 5,
        'note_left': note_x0 - 5,
        'note_right': min(note_right, cy_left - 5),
        'cy_left': cy_left - 10,
        'cy_right': mid_cy_py,
        'py_left': mid_cy_py,
        'py_right': py_right + 20,
    }


def extract_rows_from_page(page, col_bounds=None):
    """Extract structured rows using column boundaries."""
    words = page.extract_words(keep_blank_chars=False, x_tolerance=2, y_tolerance=3)
    if not words:
        return [], col_bounds

    # Strip sidebar/watermark words FIRST so they don't poison column
    # detection (Ambuja's "Creating / Value / Statutory / Reports" strip
    # at x=568 used to hijack the year-header anchor loop).
    words = _strip_sidebar_watermarks(words, page.width)

    if col_bounds is None:
        col_bounds = find_column_boundaries(words, page.width)
    if col_bounds is None:
        return [], None

    y_tol = 4
    rows_by_y = defaultdict(list)
    for w in words:
        y_key = round(w['top'] / y_tol) * y_tol
        rows_by_y[y_key].append(w)
    
    result = []
    for y in sorted(rows_by_y.keys()):
        row_words = sorted(rows_by_y[y], key=lambda w: w['x0'])
        
        parts = {'particulars': [], 'note': [], 'cy': [], 'py': []}
        
        for w in row_words:
            txt = w['text'].strip()
            if not txt:
                continue
            x0, x1 = w['x0'], w['x1']
            x_mid = (x0 + x1) / 2
            
            # Assign to column
            if x1 <= col_bounds['particulars_right'] + 5:
                parts['particulars'].append(txt)
            elif x0 >= col_bounds['note_left'] and x1 <= col_bounds['note_right'] + 5:
                if re.match(r'^\d{1,2}(-\d{1,2})?$', txt):
                    parts['note'].append(txt)
                elif x_mid < col_bounds['cy_left']:
                    parts['note'].append(txt)
                else:
                    parts['cy'].append(txt)
            elif x_mid < col_bounds['cy_right'] and x0 >= col_bounds['cy_left'] - 10:
                parts['cy'].append(txt)
            elif x0 >= col_bounds['py_left'] - 10:
                parts['py'].append(txt)
            else:
                # Ambiguous — use proximity
                dist_note = abs(x_mid - (col_bounds['note_left'] + col_bounds['note_right']) / 2)
                dist_cy = abs(x_mid - (col_bounds['cy_left'] + col_bounds['cy_right']) / 2)
                dist_py = abs(x_mid - (col_bounds['py_left'] + col_bounds['py_right']) / 2)
                
                min_dist = min(dist_note, dist_cy, dist_py)
                if min_dist == dist_py:
                    parts['py'].append(txt)
                elif min_dist == dist_cy:
                    parts['cy'].append(txt)
                elif re.match(r'^\d{1,2}$', txt):
                    parts['note'].append(txt)
                else:
                    parts['particulars'].append(txt)
        
        particulars = ' '.join(parts['particulars']).strip()
        note = ' '.join(parts['note']).strip()
        cy_raw = ' '.join(parts['cy']).strip()
        py_raw = ' '.join(parts['py']).strip()
        
        cy_val = parse_value(cy_raw)
        py_val = parse_value(py_raw)
        
        if not particulars and not cy_raw and not py_raw:
            continue
        
        result.append({
            'particulars': particulars,
            'note': note,
            'cy': cy_val,
            'py': py_val,
            'cy_raw': cy_raw,
            'py_raw': py_raw,
        })
    
    return result, col_bounds


def parse_value(s):
    s = s.strip()
    if not s:
        return ''
    if s == '-':
        return '-'
    s_clean = s.replace(' ', '')
    m = re.match(r'^\(([\d,]+\.?\d*)\)$', s_clean)
    if m:
        return -float(m.group(1).replace(',', ''))
    m = re.match(r'^([\d,]+\.?\d*)$', s_clean)
    if m:
        return float(m.group(1).replace(',', ''))
    return s


def is_stop_line(text):
    low = text.lower().strip()
    return any(p in low for p in STOP_PHRASES)


def detect_col_headers(rows):
    for r in rows[:15]:
        for field in ['particulars', 'cy_raw', 'py_raw']:
            text = r.get(field, '')
            dates = re.findall(r'(?:31\s*March\s*\d{4}|March\s*31(?:st)?,?\s*\d{4})', text, re.I)
            if len(dates) >= 2:
                return dates[0].strip(), dates[1].strip()
    # Check combined text
    for r in rows[:15]:
        text = ' '.join([r.get('particulars',''), r.get('cy_raw',''), r.get('py_raw','')])
        dates = re.findall(r'(?:31\s*March\s*\d{4}|March\s*31(?:st)?,?\s*\d{4})', text, re.I)
        if len(dates) >= 2:
            return dates[0].strip(), dates[1].strip()
    return 'Current Year', 'Prior Year'


def filter_data_rows(rows):
    data_starts = ['income', 'revenue', 'assets', 'equity and liabilities',
                   'shareholder', 'expenses', 'non-current', 'non- current',
                   'current assets', 'equity and liabilit']
    
    start = 0
    for i, r in enumerate(rows):
        low = r['particulars'].lower().strip()
        if any(k in low for k in ['particulars', 'all amount', 'all figures', 'statement of',
                                    'balance sheet as at', 'profit and loss', 'profit & loss',
                                    'indian rupees', 'rupees millions', 'lakhs inr',
                                    'for the year', 'as at as at', 'ended ended']):
            start = i + 1
            continue
        if re.match(r'^(as at|for the|ended|note)', low):
            start = i + 1
            continue
        if re.match(r'^\d{1,2}\s+(march|april|may)', low, re.I):
            start = i + 1
            continue
        if any(low.startswith(m) or low == m for m in data_starts):
            start = i
            break
        if isinstance(r['cy'], (int, float)) and r['particulars']:
            start = i
            break
    
    end = len(rows)
    for i in range(start, len(rows)):
        if is_stop_line(rows[i]['particulars']):
            end = i
            break
    
    filtered = []
    for r in rows[start:end]:
        part = r['particulars'].strip()
        low = part.lower()
        cy = r['cy']
        py = r['py']

        # Skip pure page numbers (bare 1-4 digit numbers)
        if re.match(r'^\d{1,4}$', part) and cy == '' and py == '':
            continue
        # Skip CIN/ICAI rows
        if re.match(r'^(CIN|ICAI)', part, re.I):
            continue
        # Skip "Annual Report" headers bleeding into data
        if 'annual report' in low:
            continue
        # Skip mid-table "PARTICULARS NOTE" header rows
        if low.startswith('particulars') and ('note' in low or 'schedule' in low):
            continue
        # Skip rows where value looks like a date (312025, 31032025, 31032024)
        if isinstance(cy, (int, float)) and cy > 100000 and re.match(r'^3[01]0?[3-9]20\d{2}$', str(int(cy))):
            continue
        if isinstance(py, (int, float)) and py > 100000 and re.match(r'^3[01]0?[3-9]20\d{2}$', str(int(py))):
            continue
        # Skip bare year values (2024, 2025) as data
        if isinstance(cy, (int, float)) and 2020 <= cy <= 2030 and (py == '' or py == 0):
            continue
        if isinstance(cy, (int, float)) and cy == 2025.0:
            continue
        # Skip company name / header rows (no label but large bare number)
        if not part and isinstance(cy, (int, float)) and cy > 10000 and py == '':
            continue
        # Skip rows that are just "for the year ended" or "as at" mid-table
        if re.match(r'^(for the year|as at|ended|note\s*no|rupees|lakhs|crore)', low):
            continue
        # Skip OCR-garbled rows (repeated characters: "ttaaxx", "ccuurr", "eett")
        if re.search(r'(.)\1{2,}', low) and len(re.findall(r'(.)\1{2,}', low)) >= 2:
            continue
        # Skip company address / registered office rows
        if re.search(r'(regd|registered)\s+office|legislative\s+assembl|cin\s*:', low):
            continue
        # Skip page-design watermark text (Ambuja-style): rows where every
        # column is empty except one that contains a single design token like
        # "Creating", "Value", "Statutory", "Reports", etc.
        cy_str = (r.get('cy_raw') or '').strip().lower()
        py_str = (r.get('py_raw') or '').strip().lower()
        note_str = (r.get('note') or '').strip().lower()
        non_empty_strs = [s for s in (part.lower(), note_str, cy_str, py_str) if s]
        if len(non_empty_strs) == 1:
            tokens = non_empty_strs[0].split()
            if len(tokens) <= 2 and all(t in WATERMARK_TOKENS for t in tokens):
                continue

        filtered.append(r)

    # ── Column concatenation detection ──
    # If a value is > 100x the median, it's likely two columns merged (e.g., "4444248399")
    all_cy = [abs(r['cy']) for r in filtered if isinstance(r['cy'], (int, float)) and r['cy'] != 0]
    all_py = [abs(r['py']) for r in filtered if isinstance(r['py'], (int, float)) and r['py'] != 0]
    
    if all_cy:
        median_cy = sorted(all_cy)[len(all_cy) // 2]
        if median_cy > 0:
            for r in filtered:
                if isinstance(r['cy'], (int, float)) and median_cy > 0:
                    if abs(r['cy']) > median_cy * 200 and abs(r['cy']) > 1000000:
                        # Flag as suspect — set to 0 and mark in label
                        r['_warning'] = f"SUSPECT VALUE CY={r['cy']} (>200x median {median_cy:.0f})"
                        r['cy'] = 0
    
    if all_py:
        median_py = sorted(all_py)[len(all_py) // 2]
        if median_py > 0:
            for r in filtered:
                if isinstance(r['py'], (int, float)) and median_py > 0:
                    if abs(r['py']) > median_py * 200 and abs(r['py']) > 1000000:
                        r['_warning'] = r.get('_warning', '') + f" SUSPECT VALUE PY={r['py']} (>200x median {median_py:.0f})"
                        r['py'] = 0
    
    return filtered


def detect_company_name(text):
    for line in text.split('\n')[:10]:
        line = line.strip()
        up = line.upper()
        if 'LIMITED' in up and all(k not in up for k in ['REGISTRATION', 'CIN', 'ICAI']):
            for rem in ['CONSOLIDATED STATEMENT OF PROFIT AND LOSS',
                        'STANDALONE STATEMENT OF PROFIT AND LOSS',
                        'CONSOLIDATED BALANCE SHEET', 'STANDALONE BALANCE SHEET',
                        'STATEMENT OF PROFIT AND LOSS', 'STATEMENT OF PROFIT & LOSS',
                        'BALANCE SHEET', 'AS AT']:
                up = up.replace(rem, '').strip()
            name = up.strip(' -')
            if name:
                return name
    for line in text.split('\n')[:5]:
        line = line.strip()
        if line and len(line) > 5 and all(k not in line.upper() for k in
                                           ['BALANCE', 'PROFIT', 'STATEMENT', 'ALL AMOUNT', 'CIN', '(']):
            return line
    return ''


def _is_dual_column_page(page):
    """Detect if a page has BS and PL side-by-side (landscape dual-column layout)."""
    if page.width <= page.height:
        return False  # Portrait — not dual column

    text = (page.extract_text() or '').upper()
    has_bs = any(k in text for k in ['BALANCE SHEET', 'TOTAL ASSETS', 'EQUITY AND LIABILITIES'])
    has_pl = any(k in text for k in ['REVENUE FROM OPERATIONS', 'PROFIT BEFORE TAX',
                                      'STATEMENT OF PROFIT AND LOSS', 'PROFIT AND LOSS'])
    if not (has_bs and has_pl):
        return False

    # Verify keywords are on different sides of the page
    words = page.extract_words()
    mid = page.width / 2
    bs_left = any(w['x0'] < mid and any(k in w['text'].upper() for k in ['ASSETS', 'LIABILITIES'])
                  for w in words)
    pl_right = any(w['x0'] > mid and any(k in w['text'].upper() for k in ['REVENUE', 'EXPENSE', 'PROFIT'])
                   for w in words)
    return bs_left and pl_right


def _split_dual_column(page):
    """Split a dual-column page into left (BS) and right (PL) halves.
    Returns wrapped pages with shifted coordinates so column detection works."""
    mid = page.width / 2

    left = page.crop((0, 0, mid + 5, page.height))
    right_raw = page.crop((mid - 5, 0, page.width, page.height))

    # pdfplumber crop doesn't shift coordinates — words keep original x positions.
    # Wrap the right half to shift x-coordinates so column detection works.
    right = _ShiftedPage(right_raw, x_offset=mid - 5)
    return left, right


class _ShiftedPage:
    """Wrapper around a cropped pdfplumber page that shifts x-coordinates."""
    def __init__(self, page, x_offset):
        self._page = page
        self._offset = x_offset
        self.width = page.width
        self.height = page.height

    def extract_words(self, **kwargs):
        words = self._page.extract_words(**kwargs)
        shifted = []
        for w in words:
            sw = dict(w)
            sw['x0'] = w['x0'] - self._offset
            sw['x1'] = w['x1'] - self._offset
            shifted.append(sw)
        return shifted

    def extract_text(self, **kwargs):
        return self._page.extract_text(**kwargs)


def process_pdf(pdf_path):
    pdf = pdfplumber.open(pdf_path)
    first_text = pdf.pages[0].extract_text() or ''
    company = detect_company_name(first_text)

    # ── Pre-process: expand dual-column pages into separate virtual pages ──
    pages = []
    for i, page in enumerate(pdf.pages):
        if _is_dual_column_page(page):
            left, right = _split_dual_column(page)
            pages.append(('split_left', left, i))
            pages.append(('split_right', right, i))
        else:
            pages.append(('normal', page, i))
    
    statements = []
    i = 0
    while i < len(pages):
        ptype, page, orig_idx = pages[i]
        page_text = page.extract_text() or ''
        stype = classify_page(page_text)
        
        if stype:
            page_list = [i]
            # Only look for continuation on non-split pages
            if ptype == 'normal':
                j = i + 1
                while j < len(pages):
                    _, next_page, _ = pages[j]
                    next_text = next_page.extract_text() or ''
                    next_type = classify_page(next_text)
                    if next_type is None and any(kw in next_text.upper() for kw in
                                                  ['TOTAL', 'TRADE PAYABLE', 'PROVISIONS',
                                                   'TOTAL EQUITY', 'TOTAL ASSETS', 'TOTAL LIABILITIES']):
                        page_list.append(j)
                        j += 1
                    else:
                        break
                statements.append((stype, page_list))
                i = j
            else:
                statements.append((stype, page_list))
                i += 1
        else:
            upper = page_text[:1000].upper()
            if 'BALANCE SHEET' in upper:
                statements.append(('generic_bs', [i]))
            elif any(k in upper for k in ['PROFIT AND LOSS', 'PROFIT & LOSS']):
                statements.append(('generic_pl', [i]))
            i += 1
    
    sheets = []
    for stype, page_indices in statements:
        sheet_name = STMT_NAMES.get(stype, stype)
        all_rows = []
        col_bounds = None
        # Concatenate text of all pages backing this statement to detect
        # the declared unit ("Rs in Crore" etc.). The header sits on the
        # first page only, but we scan all backers in case the first page
        # has poor OCR.
        page_text_for_unit = ""

        for pi in page_indices:
            _, page, _ = pages[pi]
            page_rows, new_bounds = extract_rows_from_page(page, col_bounds)
            if new_bounds:
                col_bounds = new_bounds
            all_rows.extend(page_rows)
            try:
                page_text_for_unit += "\n" + (page.extract_text() or "")
            except Exception:
                pass

        if not all_rows:
            continue

        # Detect unit per-statement (a single PDF can have BS in Crore and
        # P&L in Lakh — rare but observed). Normalize values to LAKHS so
        # all downstream code (mapper, per-company workbook, Airtable)
        # speaks one canonical scale.
        source_unit = detect_unit_from_text(page_text_for_unit)
        col1, col2 = detect_col_headers(all_rows)
        data_rows = filter_data_rows(all_rows)
        data_rows, factor = normalize_rows_to_lakhs(data_rows, source_unit)
        if factor != 1.0:
            print(f"  [unit] {sheet_name}: detected '{source_unit}' "
                   f"→ multiplied values by {factor:g} to normalize to Lakhs")

        if data_rows:
            sheets.append((sheet_name, data_rows, col1, col2, company,
                            source_unit))

    pdf.close()
    return sheets


def write_sheet(ws, rows, col1_header, col2_header, company_name='', statement_title='',
                source_unit='lakh'):
    header_fill = PatternFill('solid', fgColor='2F5496')
    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    title_font = Font(name='Arial', bold=True, size=13)
    sub_font = Font(name='Arial', bold=True, size=10, color='555555')
    data_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    sec_font = Font(name='Arial', size=10, bold=True, color='2F5496')
    total_bdr = Border(top=Side(style='thin', color='2F5496'), bottom=Side(style='double', color='2F5496'))
    sub_bdr = Border(bottom=Side(style='thin', color='AAAAAA'))
    
    ws.column_dimensions['A'].width = 62
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    row = 1
    if company_name:
        ws.cell(row=row, column=1, value=company_name).font = title_font
        row += 1
    if statement_title:
        # Append the unit hint so downstream consumers (and humans) can
        # tell at a glance whether normalization happened. The mapper does
        # not parse this; it always reads numeric cells as Lakhs.
        title_with_unit = statement_title
        if source_unit and source_unit != 'lakh':
            title_with_unit = (f"{statement_title}  "
                                f"(source: {source_unit.title()}; "
                                f"normalized to Lakhs)")
        elif source_unit == 'lakh':
            title_with_unit = f"{statement_title}  (Lakhs)"
        ws.cell(row=row, column=1, value=title_with_unit).font = sub_font
        row += 1
    row += 1
    
    for ci, h in enumerate(['Particulars', 'Note', col1_header, col2_header], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center' if ci > 1 else 'left', wrap_text=True)
    row += 1
    
    bold_kw = ['total', 'profit before', 'profit after', 'profit for the',
               'profit/(loss)', 'total income', 'total expense', 'total assets',
               'total equity', 'total liabilities', 'total comprehensive', 'earnings per']
    
    for r in rows:
        part = r['particulars']
        note = r['note']
        cy = r['cy']
        py = r['py']
        
        is_total = any(k in part.lower() for k in bold_kw)
        is_section = (cy == '' and py == '' and part and not re.search(r'\d', str(part)))
        is_subtotal = (not part and (isinstance(cy, (int, float)) or isinstance(py, (int, float))))
        
        font = bold_font if (is_total or is_subtotal) else (sec_font if is_section else data_font)
        
        ws.cell(row=row, column=1, value=part).font = font
        if note:
            c = ws.cell(row=row, column=2, value=note)
            c.font = data_font
            c.alignment = Alignment(horizontal='center')
        
        for vi, val in enumerate([cy, py]):
            col = 3 + vi
            cell = ws.cell(row=row, column=col)
            if isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = '#,##0.00'
            elif val == '-':
                cell.value = '-'
            elif val:
                cell.value = val
            cell.font = font
            cell.alignment = Alignment(horizontal='right')
        
        if is_total:
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = total_bdr
        elif is_subtotal:
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = sub_bdr
        row += 1
    
    ws.freeze_panes = 'A5'


def extract_tables(pdf_path, output_dir=None):
    """Extract BS/PL tables from a PDF and save as Excel."""
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"File not found: {pdf_path}")
        return

    if output_dir is None:
        output_dir = pdf_path.parent
    else:
        output_dir = Path(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\nProcessing: {pdf_path.name}")
    print("-" * 60)
    try:
        sheets = process_pdf(str(pdf_path))
        if sheets:
            wb = Workbook()
            wb.remove(wb.active)
            used = set()
            for entry in sheets:
                # Backwards-compat: process_pdf used to return 5-tuples
                # (name, rows, c1, c2, company); now returns 6-tuples
                # with source_unit appended.
                if len(entry) == 6:
                    name, rows, c1, c2, company, source_unit = entry
                else:
                    name, rows, c1, c2, company = entry
                    source_unit = "lakh"
                sn = name[:31]
                cnt = 2
                while sn in used:
                    sn = f"{name[:28]}_{cnt}"
                    cnt += 1
                used.add(sn)
                ws = wb.create_sheet(title=sn)
                write_sheet(ws, rows, c1, c2, company, name,
                            source_unit=source_unit)
                print(f"  '{sn}': {len(rows)} rows")

            out_name = pdf_path.stem + '_extracted.xlsx'
            out_path = output_dir / out_name
            _assert_not_protected(out_path)
            wb.save(str(out_path))
            print(f"  -> {out_path}")
        else:
            print(f"  No statements found")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage:")
        print("  python extract_tables.py 1_BS_PL.pdf")
        print("  python extract_tables.py 1_BS_PL.pdf 2_BS_PL.pdf")
        print("  python extract_tables.py 1_BS_PL.pdf --output ./results")
        sys.exit(1)

    out_dir = None
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            out_dir = sys.argv[idx + 1]
        pdf_files = [a for a in sys.argv[1:] if a not in ("--output", out_dir)]
    else:
        pdf_files = sys.argv[1:]

    for f in pdf_files:
        extract_tables(f, out_dir)