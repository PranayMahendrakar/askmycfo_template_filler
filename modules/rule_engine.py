"""
rule_engine.py — applies the user's rule set to an extracted "as-is" Excel file
and produces a dict of {template_label: (current_year_value, prior_year_value)}.

Pipeline position:
    PDF -> page_detector -> extract_tables -> [THIS MODULE] -> template_writer

How it works:
  1. extract_tables wrote a workbook with one sheet per statement variant.
     We pick the STANDALONE sheets only (Standalone BS / Standalone P&L);
     if no Standalone sheets exist we fall back to generic BS / P&L.
  2. We walk each row, tracking the current BS section header so bare labels
     like "Investments" can be disambiguated by section context.
  3. For each rule, we evaluate its operand list left-to-right against the
     row stream and emit a (cur, pri) pair into the result dict.

The rule shape (see data/rules.json):
    {
      "Accounts Payable": {
        "section": "Current Liabilities",
        "operands": [
          {"pattern": "trade_payables_msme", "op": "+"},
          {"pattern": "trade_payables_others", "op": "+"},
          ...
        ]
      },
      ...
    }

The operator on the first operand is the SEED — it sets the sign of that
first value (use "+" or "-"; "*" / "/" on the seed are treated as "+").
Subsequent operators apply between the accumulator and the next matched
value.

Match strategy (intentionally simple, easy to debug):
  - Normalise label and alias to lowercase, strip non-word chars to spaces.
  - For each pattern, check every alias against every row label using
    substring containment (longest matching alias wins).
  - When a pattern has a section_filter, only rows in that section are
    eligible to match it. This keeps "Investments" in Current Assets from
    being picked up by the Non-Current rule and vice versa.
  - If multiple rows match the same pattern we SUM them — handles the
    common Schedule-III pattern where "Trade Payables" is split into
    MSME / Others sub-rows.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


# ─── Section detection ──────────────────────────────────────────────

# Order matters: more specific phrases first so "non-current assets" doesn't
# get swallowed by "current assets".
_SECTION_HEADERS = [
    ("noncurrent_asset",  ["non-current assets", "non current assets", "non- current assets"]),
    ("current_asset",     ["current assets"]),
    ("noncurrent_liab",   ["non-current liabilities", "non current liabilities", "non- current liabilities"]),
    ("current_liab",      ["current liabilities"]),
    ("equity",            ["equity and liabilities", "shareholder's fund", "shareholders fund", "shareholders' fund"]),
]

# Section totals are also useful — once we see them, we know the section is
# closing so a header-less subsequent block doesn't inherit the wrong section.
_SECTION_CLOSE = [
    "total non-current assets", "total non current assets",
    "total current assets",
    "total non-current liabilities", "total non current liabilities",
    "total current liabilities",
    "total equity",
]


def _norm(s: str) -> str:
    """Lowercase + strip punctuation to single-spaces."""
    s = (s or "").lower().strip()
    s = re.sub(r"[\s\-_/&,\.\(\)\[\]:;]+", " ", s)
    return s.strip()


def _is_section_header_row(label: str, cur: Any, pri: Any) -> Optional[str]:
    """Return section key if this row is a section header (label only, no
    numeric values), else None."""
    if isinstance(cur, (int, float)) and cur != 0:
        return None
    if isinstance(pri, (int, float)) and pri != 0:
        return None
    n = _norm(label)
    if not n:
        return None
    for key, headers in _SECTION_HEADERS:
        for h in headers:
            if _norm(h) == n or _norm(h) in n:
                return key
    return None


def _is_section_close(label: str) -> bool:
    n = _norm(label)
    return any(_norm(c) in n for c in _SECTION_CLOSE)


# ─── Reading the extracted xlsx ─────────────────────────────────────

def _clean_value(v: Any) -> float:
    """Coerce a cell value into a float. Empty / dash / nil → 0.0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-", "–", "—", "nil", "Nil", "NIL"):
        return 0.0
    # Parens = negative
    if s.startswith("(") and s.endswith(")"):
        try:
            return -float(s[1:-1])
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _pick_sheets(wb: openpyxl.Workbook,
                  preference: str = "standalone") -> List[Tuple[str, str]]:
    """Return [(kind, sheet_name)] where kind in {'bs','pl'}.

    preference:
        'standalone'   -> prefer Standalone, fall back to generic
        'consolidated' -> prefer Consolidated, fall back to generic
        'auto'         -> first Standalone, then Consolidated, then generic
    """
    sheets = wb.sheetnames
    out: List[Tuple[str, str]] = []
    used: set = set()

    def _add_from(variant: Optional[str]):
        """variant in {'standalone', 'consolidated', None=generic}."""
        for sn in sheets:
            low = sn.lower()
            has_cons = "consol" in low
            has_std  = "standalone" in low
            if variant == "standalone" and not has_std:
                continue
            if variant == "consolidated" and not has_cons:
                continue
            if variant is None and (has_std or has_cons):
                continue
            is_bs = "balance" in low or low.endswith(" bs") or low == "bs"
            is_pl = ("p&l" in low or "profit" in low
                     or low.endswith(" pl") or low == "pl")
            if is_bs and "bs" not in used:
                out.append(("bs", sn)); used.add("bs")
            elif is_pl and "pl" not in used:
                out.append(("pl", sn)); used.add("pl")

    pref = (preference or "standalone").lower()
    if pref == "consolidated":
        _add_from("consolidated")
        _add_from("standalone")   # fallback if consolidated not present
        _add_from(None)           # last-resort generic
    elif pref == "auto":
        _add_from("standalone")
        _add_from("consolidated")
        _add_from(None)
    else:  # 'standalone' (default)
        _add_from("standalone")
        _add_from(None)

    return out


# Backward-compat alias for any existing callers
_pick_standalone_sheets = _pick_sheets


def read_rows(xlsx_path: str | Path,
              sheet_preference: str = "standalone") -> List[Dict[str, Any]]:
    """Read an extract_tables-produced xlsx and return a flat list of rows
    annotated with section context.

    Each row: {label, cur, pri, section, sheet_kind}
      sheet_kind in {'bs','pl'} — pl rows don't get a meaningful section.

    sheet_preference: 'standalone' | 'consolidated' | 'auto'
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    picks = _pick_sheets(wb, sheet_preference)
    rows: List[Dict[str, Any]] = []

    for kind, sn in picks:
        ws = wb[sn]
        section: Optional[str] = None
        # extract_tables lays the data out as:
        #   col A = Particulars
        #   col B = Note number
        #   col C = Current Year value
        #   col D = Prior Year value
        # The first 2-3 rows are title/subtitle/header. We skip up to the
        # row labelled "Particulars" (case-insensitive) and start after it.
        header_seen = False
        for r in ws.iter_rows(min_row=1, max_col=4, values_only=True):
            label = (r[0] or "")
            if isinstance(label, str):
                label = label.strip()
            else:
                label = str(label).strip() if label is not None else ""

            if not header_seen:
                if "particulars" in label.lower():
                    header_seen = True
                continue
            if not label or label.lower() == "none":
                continue

            cur = _clean_value(r[2] if len(r) > 2 else None)
            pri = _clean_value(r[3] if len(r) > 3 else None)

            # Section header? (label-only row with no values)
            sec = _is_section_header_row(label, cur, pri)
            if sec:
                section = sec
                continue
            if _is_section_close(label):
                # Don't reset to None — next row could still belong here.
                # Just don't carry forward into a different statement block.
                pass

            rows.append({
                "label": label,
                "cur": cur,
                "pri": pri,
                "section": section if kind == "bs" else None,
                "sheet_kind": kind,
            })
    return rows


# ─── Pattern matching ────────────────────────────────────────────────

def _row_matches_pattern(row: Dict[str, Any], pattern: Dict[str, Any]) -> Tuple[bool, int]:
    """Return (matched, score). Score = length of longest matching alias.

    Section filter is *advisory* when row.section is None — i.e. when our
    extractor couldn't determine which Balance Sheet section the row sits
    under (common for nested sub-rows like 'i) Investment in Associate
    and Joint Ventures' where the parent header carries the section, not
    the sub-row itself). Only a KNOWN-disagreeing section blocks the match.
    The score bonus in _best_pattern_for_row still favours patterns whose
    section_filter explicitly matches the row's known section, so the
    tie-breaker is preserved.
    """
    sec_filter = pattern.get("section_filter")
    row_section = row.get("section")
    if sec_filter and row_section and row_section != sec_filter:
        return False, 0

    nlabel = _norm(row["label"])
    if not nlabel:
        return False, 0

    best = 0
    for alias in pattern.get("aliases", []):
        na = _norm(alias)
        if not na:
            continue
        if na in nlabel:
            best = max(best, len(na))
    return (best > 0), best


def _is_total_row(label: str) -> bool:
    """Return True for rows that are sums/totals — they're aggregates of
    other rows, not source data, and matching them double-counts. Also
    catches the 'Profit before tax' / 'Profit for the year' P&L sub-totals.
    """
    n = _norm(label)
    if not n:
        return True
    if n.startswith("total ") or n.startswith("sub total") or n.startswith("subtotal"):
        return True
    if n.startswith("profit for the") or n.startswith("profit before tax") \
       or n.startswith("loss for the") or n.startswith("loss before tax"):
        return True
    if "earnings per share" in n or "earning per share" in n:
        return True
    return False


def _best_pattern_for_row(row: Dict[str, Any], patterns: Dict[str, Dict]) -> Optional[str]:
    """Pick THE single pattern with the longest matching alias for this row.
    Returns the pattern id, or None if nothing matched."""
    # Totals / sub-totals aren't source data — skip them so they don't get
    # double-counted into a pattern bucket.
    if _is_total_row(row["label"]):
        return None
    best_pid = None
    best_score = 0
    for pid, p in patterns.items():
        ok, score = _row_matches_pattern(row, p)
        # Bonus 5 for patterns with a section_filter that DID match —
        # tie-breaker in favour of the more-specific pattern.
        if ok and p.get("section_filter") and row.get("section") == p["section_filter"]:
            score += 5
        if ok and score > best_score:
            best_score = score
            best_pid = pid
    return best_pid


def bucket_rows_by_pattern(rows: List[Dict[str, Any]],
                            patterns: Dict[str, Dict]) -> Dict[str, Dict[str, Any]]:
    """For each pattern, sum the cy/pri of every row that best-matches it.
    Returns {pattern_id: {'cur': float, 'pri': float, 'sources': [...]}}."""
    buckets: Dict[str, Dict[str, Any]] = {
        pid: {"cur": 0.0, "pri": 0.0, "sources": []} for pid in patterns
    }
    for row in rows:
        pid = _best_pattern_for_row(row, patterns)
        if not pid:
            continue
        b = buckets[pid]
        b["cur"] += row["cur"]
        b["pri"] += row["pri"]
        b["sources"].append({
            "label": row["label"], "cur": row["cur"], "pri": row["pri"],
            "section": row.get("section"),
        })
    return buckets


# ─── Rule evaluation ─────────────────────────────────────────────────

def _apply_op(acc: float, val: float, op: str) -> float:
    if op == "+":
        return acc + val
    if op == "-":
        return acc - val
    if op == "*":
        return acc * val
    if op == "/":
        return acc / val if val != 0 else 0.0
    return acc + val


def _seed_value(val: float, op: str) -> float:
    """First operand: only sign matters. '-' inverts, everything else keeps
    it as-is. '*' / '/' on the seed make no sense — treat as '+'."""
    if op == "-":
        return -val
    return val


def evaluate_rules(rules: Dict[str, Dict],
                   patterns: Dict[str, Dict],
                   buckets: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Apply every rule. Returns:
        {template_label: {'cur': float, 'pri': float, 'trace': [...]}}.
    `trace` is a per-operand record of (pattern, op, matched_cur, matched_pri,
    source_labels) so the UI / debug log can show exactly how a number was
    built up."""
    out: Dict[str, Dict[str, Any]] = {}
    for label, rule in rules.items():
        if label.startswith("_"):  # skip _comment etc.
            continue
        operands = rule.get("operands") or []
        cur_acc: Optional[float] = None
        pri_acc: Optional[float] = None
        trace: List[Dict[str, Any]] = []

        for i, op_entry in enumerate(operands):
            pid = op_entry.get("pattern")
            op = (op_entry.get("op") or "+").strip()
            bucket = buckets.get(pid) or {"cur": 0.0, "pri": 0.0, "sources": []}
            mc, mp = bucket["cur"], bucket["pri"]
            srcs = bucket.get("sources", [])

            if i == 0:
                cur_acc = _seed_value(mc, op)
                pri_acc = _seed_value(mp, op)
            else:
                cur_acc = _apply_op(cur_acc or 0.0, mc, op)
                pri_acc = _apply_op(pri_acc or 0.0, mp, op)

            trace.append({
                "pattern": pid,
                "pattern_name": (patterns.get(pid) or {}).get("name", pid),
                "op": op,
                "matched_cur": mc,
                "matched_pri": mp,
                "n_sources": len(srcs),
                "sources": srcs[:6],  # cap so the report stays small
            })

        out[label] = {
            "cur": round(cur_acc or 0.0, 2),
            "pri": round(pri_acc or 0.0, 2),
            "trace": trace,
        }
    return out


# ─── Convenience entry point ─────────────────────────────────────────

def run(extracted_xlsx: str | Path,
        rules_path: str | Path,
        patterns_path: str | Path,
        sheet_preference: str = "standalone") -> Dict[str, Any]:
    """One-call helper used by app.py. Returns:
        {
          'values':   {label: {cur, pri, trace}},
          'rows':     [list of all extracted rows we considered],
          'buckets':  {pattern_id: {cur, pri, sources}}  -- for debugging,
          'unmatched':[rows that didn't match any pattern],
        }

    sheet_preference: 'standalone' | 'consolidated' | 'auto'
    """
    with open(rules_path, "r", encoding="utf-8") as f:
        rules = json.load(f)
    with open(patterns_path, "r", encoding="utf-8") as f:
        patterns_doc = json.load(f)
    patterns = patterns_doc.get("patterns", {})

    rows = read_rows(extracted_xlsx, sheet_preference=sheet_preference)
    buckets = bucket_rows_by_pattern(rows, patterns)
    values = evaluate_rules(rules, patterns, buckets)

    unmatched = []
    for row in rows:
        if not _best_pattern_for_row(row, patterns):
            # Skip rows that are clearly headers / totals / page noise — they
            # legitimately don't match and aren't useful as warnings.
            n = _norm(row["label"])
            if (not n) or _is_total_row(row["label"]) \
               or "particulars" in n or "annual report" in n:
                continue
            if row["cur"] == 0 and row["pri"] == 0:
                continue
            unmatched.append(row)

    return {
        "values": values,
        "rows": rows,
        "buckets": buckets,
        "unmatched": unmatched,
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python rule_engine.py <extracted.xlsx>")
        sys.exit(1)
    base = Path(__file__).parent.parent / "data"
    result = run(sys.argv[1], base / "rules.json", base / "patterns.json")
    print(f"Rows read: {len(result['rows'])}")
    print(f"Unmatched: {len(result['unmatched'])}")
    print()
    for label, v in result["values"].items():
        print(f"  {label:42s} CY={v['cur']:>14,.2f}  PY={v['pri']:>14,.2f}")
