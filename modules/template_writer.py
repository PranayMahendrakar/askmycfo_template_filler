"""
template_writer.py — fills Input_Templates.xlsx with values computed by
rule_engine.evaluate_rules().

What it does:
  - Loads the master Input_Templates.xlsx (read-only on disk; never overwritten).
  - Writes the rule-engine values into column F (CY) and column G (PY).
  - Writes the formula rows (Total Networth, Total Assets, DIFFERENCE, etc.)
    using Excel formulas — never hardcoded sums — so the spreadsheet stays
    dynamic if the user edits any number.
  - Adds row 1 with the company name and row 2 with the CY/PY headers.
  - Saves under output_path. The master template is never modified.

The labels and rows come from data/cell_map.json:
  data_rows    — labels that map to a single row that takes a value
  formula_rows — labels with an Excel formula (no rule needed)
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Cell styling — kept consistent with bs_pl_mapper.build_report so reports
# look like they came from the same family.
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10, color="0000FF")
TOT_FILL  = PatternFill("solid", fgColor="D9D9D9")
TOT_FONT  = Font(name="Arial", bold=True, size=9, color="333333")
DIFF_FILL = PatternFill("solid", fgColor="4F46E5")
DIFF_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
NUM_FMT   = '#,##0.00;(#,##0.00);"-"'
THIN_BDR  = Border(bottom=Side(style="thin", color="EEEEEE"))


def write_filled_template(master_template: str | Path,
                          output_path: str | Path,
                          values: Dict[str, Dict[str, Any]],
                          cell_map: Dict[str, Any],
                          company_name: str = "",
                          fy_label: str = "") -> str:
    """Open the master template, write values + formulas, save to output_path.

    Args:
      master_template: path to Input_Templates.xlsx (the read-only master).
      output_path:     where to save the filled copy.
      values:          rule_engine output {template_label: {cur, pri, trace}}.
      cell_map:        parsed data/cell_map.json.
      company_name:    written into row 1 (D1 + E1 merged area).
      fy_label:        column header label (e.g. "FY 2024-25").

    Returns the absolute path of the saved file.
    """
    wb = openpyxl.load_workbook(master_template)
    ws = wb.active

    # Column setup. Labels are already in column E (rows 3-74 in the master).
    # We use F = CY, G = PY.
    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    # Row 1 — company name banner
    if company_name:
        ws["E1"] = company_name
        ws["E1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")

    # Row 2 — column headers
    cy_label = f"Current Year{(' — ' + fy_label) if fy_label else ''}"
    py_label = "Prior Year"
    ws["F2"] = cy_label
    ws["G2"] = py_label
    for c in ("F2", "G2"):
        ws[c].font = HDR_FONT
        ws[c].fill = HDR_FILL
        ws[c].alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30

    # ── Data rows ──
    for label, info in cell_map.get("data_rows", {}).items():
        row = info["row"]
        v = values.get(label) or {"cur": 0.0, "pri": 0.0}
        for col, key in (("F", "cur"), ("G", "pri")):
            cell = ws[f"{col}{row}"]
            cell.value = float(v[key]) if v[key] is not None else 0.0
            cell.font = DATA_FONT
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BDR

    # ── Formula rows ──
    for label, info in cell_map.get("formula_rows", {}).items():
        row = info["row"]
        formula_tpl = info["formula"]
        for col in ("F", "G"):
            cell = ws[f"{col}{row}"]
            cell.value = formula_tpl.replace("{c}", col)
            # DIFFERENCE row gets the special purple highlight
            if label.upper().startswith("DIFFERENCE"):
                cell.font = DIFF_FONT
                cell.fill = DIFF_FILL
            else:
                cell.font = TOT_FONT
                cell.fill = TOT_FILL
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right")

    # Freeze the labels column so scrolling keeps them visible
    ws.freeze_panes = "F3"

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out.resolve())


# Try to use the xlsx skill's recalc helper if present (it asks LibreOffice
# to evaluate formulas so users see numbers instead of "=SUM(...)" placeholders
# until they open the file). It's optional — Excel/Sheets will recalc on
# open anyway.
def try_recalc(xlsx_path: str | Path) -> bool:
    import os
    candidates = [
        "/mnt/skills/public/xlsx/scripts/recalc.py",
        str(Path(__file__).parent.parent / "scripts" / "recalc.py"),
    ]
    for c in candidates:
        if Path(c).exists():
            rc = os.system(f"python3 {c} '{xlsx_path}' > /dev/null 2>&1")
            return rc == 0
    return False


if __name__ == "__main__":
    # quick smoke test: write a template with all-zero values
    base = Path(__file__).parent.parent
    cm = json.loads((base / "data" / "cell_map.json").read_text(encoding="utf-8"))
    values = {label: {"cur": 0.0, "pri": 0.0, "trace": []}
              for label in cm.get("data_rows", {})}
    out = write_filled_template(
        base / "Input_Templates.xlsx",
        base / "outputs" / "_smoke_test.xlsx",
        values, cm,
        company_name="SMOKE TEST CO",
        fy_label="FY 2024-25",
    )
    print(f"wrote {out}")
